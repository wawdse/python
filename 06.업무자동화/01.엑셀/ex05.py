from openpyxl import load_workbook
wb = load_workbook('data/sample.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=5):
    for col in row:
        print(col.value, end=',')
    print('')