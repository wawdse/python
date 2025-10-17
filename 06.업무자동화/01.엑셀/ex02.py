from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = '기본시트'

ws = wb.create_sheet()
ws.title = '나의시트'
ws.sheet_properties.tabColor = 'FF00DD'

ws1 = wb.create_sheet('너의시트')

your_sheet = wb['너의시트']
your_sheet['A1'] = '테스트'

copy_sheet = wb.copy_worksheet(your_sheet)
copy_sheet.title = '카피시트'

print(wb.sheetnames)
wb.save('data/sample.xlsx')
wb.close()