from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title='나의시트'

ws['A1'] = 1
ws['A2'] = 2

ws['B1'] = 3
ws['B2'] = 4

print(ws['A1'])
print(ws['A1'].value)
print(ws.cell(row=1, column=1).value)

for col in range(1, 3):
    for row in range(1, 3):
        print(ws.cell(row=row, column=col).value)

wb.save('data/sample.xlsx')
wb.close()